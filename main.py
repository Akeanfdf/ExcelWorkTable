import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import json
import threading

CONFIG_FILE = "mapping_config.json"

# ─── Color palette ───────────────────────────────────────────────
BG        = "#F0F4F8"
SIDEBAR   = "#1E2A3A"
ACCENT    = "#3B82F6"
ACCENT2   = "#10B981"
DANGER    = "#EF4444"
WHITE     = "#FFFFFF"
CARD      = "#FFFFFF"
TEXT      = "#1F2937"
MUTED     = "#6B7280"
BORDER    = "#E5E7EB"

class ExcelBatchApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 批量表单生成器")
        self.geometry("1080x700")
        self.minsize(900, 600)
        self.configure(bg=BG)

        self.data_file   = tk.StringVar()
        self.tmpl_file   = tk.StringVar()
        self.output_dir  = tk.StringVar()
        self.name_col    = tk.StringVar()
        self.sheet_name  = tk.StringVar(value="Sheet1")
        self.status_var  = tk.StringVar(value="就绪")
        self.progress    = tk.DoubleVar()
        self.mapping_rows = []   # list of (col_var, cell_var) pairs
        self.df_columns   = []

        self._build_ui()
        self._load_config()

    # ─── UI construction ──────────────────────────────────────────
    def _build_ui(self):
        # Sidebar
        sidebar = tk.Frame(self, bg=SIDEBAR, width=220)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)

        tk.Label(sidebar, text="⚡ Excel 批量\n表单生成器",
                 bg=SIDEBAR, fg=WHITE,
                 font=("Arial", 14, "bold"), justify="center").pack(pady=(30, 10))
        tk.Frame(sidebar, bg="#334155", height=1).pack(fill=tk.X, padx=20, pady=10)

        self._nav_btn(sidebar, "📂  文件配置",  lambda: self._show("files"))
        self._nav_btn(sidebar, "🔗  字段映射",  lambda: self._show("mapping"))
        self._nav_btn(sidebar, "▶   开始生成",  lambda: self._show("run"))
        self._nav_btn(sidebar, "📋  使用说明",  lambda: self._show("help"))

        tk.Frame(sidebar, bg="#334155", height=1).pack(fill=tk.X, padx=20, pady=10)
        tk.Label(sidebar, text="v1.0  Excel Generator",
                 bg=SIDEBAR, fg="#64748B", font=("Arial", 9)).pack(side=tk.BOTTOM, pady=14)

        # Main area
        self.main = tk.Frame(self, bg=BG)
        self.main.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.pages = {}
        for name in ("files", "mapping", "run", "help"):
            f = tk.Frame(self.main, bg=BG)
            self.pages[name] = f

        self._build_files_page()
        self._build_mapping_page()
        self._build_run_page()
        self._build_help_page()
        self._show("files")

    def _nav_btn(self, parent, text, cmd):
        btn = tk.Button(parent, text=text, command=cmd,
                        bg=SIDEBAR, fg="#CBD5E1",
                        font=("Arial", 11), bd=0, anchor="w",
                        padx=24, pady=10, activebackground="#334155",
                        activeforeground=WHITE, cursor="hand2")
        btn.pack(fill=tk.X)

    def _show(self, name):
        for p in self.pages.values():
            p.pack_forget()
        self.pages[name].pack(fill=tk.BOTH, expand=True, padx=24, pady=20)

    def _card(self, parent, title):
        frame = tk.LabelFrame(parent, text=f"  {title}  ",
                              bg=CARD, fg=TEXT,
                              font=("Arial", 11, "bold"),
                              bd=1, relief="groove",
                              labelanchor="nw")
        return frame

    # ─── Files Page ───────────────────────────────────────────────
    def _build_files_page(self):
        p = self.pages["files"]
        tk.Label(p, text="文件配置", bg=BG, fg=TEXT,
                 font=("Arial", 18, "bold")).pack(anchor="w", pady=(0, 16))

        card = self._card(p, "选择文件")
        card.pack(fill=tk.X, pady=(0, 12))
        card.configure(padx=16, pady=12)

        rows = [
            ("📊 数据源文件 (.xlsx)", self.data_file, self._pick_data),
            ("📄 模板文件 (.xlsx)",   self.tmpl_file, self._pick_tmpl),
            ("📁 输出目录",           self.output_dir, self._pick_output),
        ]
        for label, var, cmd in rows:
            row = tk.Frame(card, bg=CARD)
            row.pack(fill=tk.X, pady=4)
            tk.Label(row, text=label, bg=CARD, fg=TEXT,
                     font=("Arial", 10), width=22, anchor="w").pack(side=tk.LEFT)
            tk.Entry(row, textvariable=var, font=("Arial", 10),
                     bg=BG, relief="flat", bd=2).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
            tk.Button(row, text="浏览", command=cmd,
                      bg=ACCENT, fg=WHITE, font=("Arial", 9, "bold"),
                      bd=0, padx=12, pady=4, cursor="hand2").pack(side=tk.LEFT)

        card2 = self._card(p, "输出设置")
        card2.pack(fill=tk.X, pady=(0, 12))
        card2.configure(padx=16, pady=12)

        r1 = tk.Frame(card2, bg=CARD); r1.pack(fill=tk.X, pady=4)
        tk.Label(r1, text="文件命名列（数据源中的列名）：", bg=CARD, fg=TEXT,
                 font=("Arial", 10), width=26, anchor="w").pack(side=tk.LEFT)
        self.name_col_combo = ttk.Combobox(r1, textvariable=self.name_col,
                                           font=("Arial", 10), width=28)
        self.name_col_combo.pack(side=tk.LEFT, padx=8)

        r2 = tk.Frame(card2, bg=CARD); r2.pack(fill=tk.X, pady=4)
        tk.Label(r2, text="模板工作表名称：", bg=CARD, fg=TEXT,
                 font=("Arial", 10), width=26, anchor="w").pack(side=tk.LEFT)
        tk.Entry(r2, textvariable=self.sheet_name, font=("Arial", 10),
                 bg=BG, relief="flat", bd=2, width=30).pack(side=tk.LEFT, padx=8)

        tk.Button(p, text="加载数据源列 →", command=self._load_columns,
                  bg=ACCENT2, fg=WHITE, font=("Arial", 11, "bold"),
                  bd=0, padx=20, pady=8, cursor="hand2").pack(anchor="e", pady=8)

    def _pick_data(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if f:
            self.data_file.set(f)
            self._load_columns()

    def _pick_tmpl(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if f: self.tmpl_file.set(f)

    def _pick_output(self):
        d = filedialog.askdirectory()
        if d: self.output_dir.set(d)

    def _load_columns(self):
        path = self.data_file.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("提示", "请先选择有效的数据源文件")
            return
        try:
            df = pd.read_excel(path, nrows=0)
            self.df_columns = list(df.columns)
            self.name_col_combo["values"] = self.df_columns
            if self.df_columns and not self.name_col.get():
                self.name_col.set(self.df_columns[0])
            self._refresh_mapping_combos()
            messagebox.showinfo("成功", f"已加载 {len(self.df_columns)} 个列名")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    # ─── Mapping Page ─────────────────────────────────────────────
    def _build_mapping_page(self):
        p = self.pages["mapping"]
        tk.Label(p, text="字段映射", bg=BG, fg=TEXT,
                 font=("Arial", 18, "bold")).pack(anchor="w", pady=(0, 6))
        tk.Label(p, text="设置「数据源列名」→「模板单元格地址」的对应关系（如 B3、D5）",
                 bg=BG, fg=MUTED, font=("Arial", 10)).pack(anchor="w", pady=(0, 12))

        # Header
        hdr = tk.Frame(p, bg=ACCENT)
        hdr.pack(fill=tk.X)
        for txt, w in [("  #", 4), ("数据源列名", 28), ("→  模板单元格", 18), ("", 8)]:
            tk.Label(hdr, text=txt, bg=ACCENT, fg=WHITE,
                     font=("Arial", 10, "bold"), width=w, anchor="w", pady=6).pack(side=tk.LEFT)

        # Scrollable area
        wrapper = tk.Frame(p, bg=BORDER, bd=1, relief="flat")
        wrapper.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(wrapper, bg=CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(wrapper, orient="vertical", command=canvas.yview)
        self.map_frame = tk.Frame(canvas, bg=CARD)
        self.map_frame.bind("<Configure>",
                            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.map_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Buttons
        btn_row = tk.Frame(p, bg=BG)
        btn_row.pack(fill=tk.X, pady=10)
        tk.Button(btn_row, text="＋ 添加映射行", command=self._add_mapping_row,
                  bg=ACCENT, fg=WHITE, font=("Arial", 10, "bold"),
                  bd=0, padx=16, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(btn_row, text="💾 保存配置", command=self._save_config,
                  bg=ACCENT2, fg=WHITE, font=("Arial", 10, "bold"),
                  bd=0, padx=16, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(btn_row, text="🗑 清空", command=self._clear_mapping,
                  bg=DANGER, fg=WHITE, font=("Arial", 10, "bold"),
                  bd=0, padx=16, pady=6, cursor="hand2").pack(side=tk.LEFT)

        # Add 3 default rows
        for _ in range(3):
            self._add_mapping_row()

    def _add_mapping_row(self, col_val="", cell_val=""):
        idx = len(self.mapping_rows) + 1
        row_frame = tk.Frame(self.map_frame, bg=CARD if idx % 2 else "#F9FAFB")
        row_frame.pack(fill=tk.X)

        tk.Label(row_frame, text=f"  {idx}", bg=row_frame["bg"], fg=MUTED,
                 font=("Arial", 10), width=4).pack(side=tk.LEFT)

        col_var  = tk.StringVar(value=col_val)
        cell_var = tk.StringVar(value=cell_val)

        col_combo = ttk.Combobox(row_frame, textvariable=col_var,
                                 values=self.df_columns, font=("Arial", 10), width=26)
        col_combo.pack(side=tk.LEFT, padx=6, pady=5)

        tk.Label(row_frame, text="→", bg=row_frame["bg"], fg=ACCENT,
                 font=("Arial", 12, "bold")).pack(side=tk.LEFT)

        cell_entry = tk.Entry(row_frame, textvariable=cell_var,
                              font=("Arial", 10), bg="#F0F4F8",
                              relief="flat", bd=2, width=14)
        cell_entry.pack(side=tk.LEFT, padx=6)

        def remove():
            self.mapping_rows = [(c, v) for c, v in self.mapping_rows
                                 if c is not col_var]
            row_frame.destroy()
            self._renumber_rows()

        tk.Button(row_frame, text="✕", command=remove,
                  bg=row_frame["bg"], fg=DANGER,
                  font=("Arial", 11), bd=0, cursor="hand2").pack(side=tk.LEFT, padx=4)

        self.mapping_rows.append((col_var, cell_var))

    def _renumber_rows(self):
        for i, child in enumerate(self.map_frame.winfo_children()):
            for widget in child.winfo_children():
                if isinstance(widget, tk.Label) and widget.cget("width") == 4:
                    widget.config(text=f"  {i+1}")
                    break

    def _refresh_mapping_combos(self):
        for child in self.map_frame.winfo_children():
            for widget in child.winfo_children():
                if isinstance(widget, ttk.Combobox):
                    widget["values"] = self.df_columns

    def _clear_mapping(self):
        if messagebox.askyesno("确认", "清空所有映射行？"):
            for w in self.map_frame.winfo_children():
                w.destroy()
            self.mapping_rows.clear()

       # ─── Run Page ─────────────────────────────────────────────────
    def _build_run_page(self):
        p = self.pages["run"]
        tk.Label(p, text="开始生成", bg=BG, fg=TEXT,
                 font=("Arial", 18, "bold")).pack(anchor="w", pady=(0, 10))

        # 底部操作区先占位，避免被「可拉伸」的日志区挤出可视区域
        action = tk.Frame(p, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        action.pack(side=tk.BOTTOM, fill=tk.X, pady=(8, 0))
        action_inner = tk.Frame(action, bg=CARD)
        action_inner.pack(fill=tk.X, padx=16, pady=12)
        tk.Label(
            action_inner,
            text="先刷新摘要核对配置，再开始批量生成（生成过程可在日志区查看）",
            bg=CARD, fg=MUTED, font=("Arial", 9), wraplength=720, justify="left",
        ).pack(anchor="w", pady=(0, 8))
        btn_row = tk.Frame(action_inner, bg=CARD)
        btn_row.pack(fill=tk.X)
        tk.Button(btn_row, text="▶  刷新摘要", command=self._refresh_summary,
                  bg=ACCENT, fg=WHITE, font=("Arial", 11),
                  bd=0, padx=18, pady=8, cursor="hand2").pack(side=tk.LEFT, padx=(0, 10))
        tk.Button(btn_row, text="🚀  开始批量生成", command=self._run_generate,
                  bg=ACCENT2, fg=WHITE, font=("Arial", 12, "bold"),
                  bd=0, padx=24, pady=8, cursor="hand2").pack(side=tk.LEFT)

        # 中间可伸缩：摘要 + 进度 + 日志（仅占按钮上方空间）
        body = tk.Frame(p, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)

        self.summary_card = self._card(body, "当前配置摘要")
        self.summary_card.pack(fill=tk.X, pady=(0, 10))
        self.summary_card.configure(padx=16, pady=12)
        self.summary_text = tk.Text(self.summary_card, height=5, font=("Courier", 10),
                                    bg="#F8FAFC", fg=TEXT, relief="flat", state="disabled")
        self.summary_text.pack(fill=tk.X)

        prog_card = self._card(body, "生成进度")
        prog_card.pack(fill=tk.X, pady=(0, 10))
        prog_card.configure(padx=16, pady=12)

        self.prog_bar = ttk.Progressbar(prog_card, variable=self.progress,
                                        maximum=100, mode="determinate")
        self.prog_bar.pack(fill=tk.X, pady=(0, 6))
        tk.Label(prog_card, textvariable=self.status_var, bg=CARD, fg=MUTED,
                 font=("Arial", 10)).pack(anchor="w")

        log_card = self._card(body, "日志")
        log_card.pack(fill=tk.BOTH, expand=True)
        log_card.configure(padx=16, pady=12)
        self.log = tk.Text(log_card, font=("Courier", 9), bg="#0F172A", fg="#94A3B8",
                           relief="flat", state="disabled")
        sb = ttk.Scrollbar(log_card, command=self.log.yview)
        self.log.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.log.pack(fill=tk.BOTH, expand=True)

        # Run button
        btn_row = tk.Frame(p, bg=BG); btn_row.pack(fill=tk.X, pady=10)
        tk.Button(btn_row, text="▶  刷新摘要", command=self._refresh_summary,
                  bg=ACCENT, fg=WHITE, font=("Arial", 11),
                  bd=0, padx=18, pady=8, cursor="hand2").pack(side=tk.LEFT, padx=(0, 10))
        tk.Button(btn_row, text="🚀  开始批量生成", command=self._run_generate,
                  bg=ACCENT2, fg=WHITE, font=("Arial", 12, "bold"),
                  bd=0, padx=24, pady=8, cursor="hand2").pack(side=tk.LEFT)

    def _refresh_summary(self):
        mapping = [(c.get(), v.get()) for c, v in self.mapping_rows
                   if c.get() and v.get()]
        txt = (
            f"数据源文件：{self.data_file.get() or '未设置'}\n"
            f"模板文件：  {self.tmpl_file.get() or '未设置'}\n"
            f"输出目录：  {self.output_dir.get() or '未设置'}\n"
            f"命名列：    {self.name_col.get() or '未设置'}\n"
            f"模板Sheet： {self.sheet_name.get()}\n"
            f"映射数量：  {len(mapping)} 条\n"
        )
        for col, cell in mapping:
            txt += f"  · {col:<20} → {cell}\n"
        self.summary_text.config(state="normal")
        self.summary_text.delete("1.0", tk.END)
        self.summary_text.insert("1.0", txt)
        self.summary_text.config(state="disabled")

    def _log(self, msg, color="#94A3B8"):
        self.log.config(state="normal")
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.log.config(state="disabled")

    def _run_generate(self):
        self._refresh_summary()
        errors = []
        if not self.data_file.get() or not os.path.exists(self.data_file.get()):
            errors.append("数据源文件无效")
        if not self.tmpl_file.get() or not os.path.exists(self.tmpl_file.get()):
            errors.append("模板文件无效")
        if not self.output_dir.get():
            errors.append("未设置输出目录")
        if not self.name_col.get():
            errors.append("未设置命名列")
        mapping = [(c.get(), v.get()) for c, v in self.mapping_rows
                   if c.get() and v.get()]
        if not mapping:
            errors.append("没有有效的字段映射")
        if errors:
            messagebox.showerror("配置错误", "\n".join(errors))
            return

        self._show("run")
        threading.Thread(target=self._do_generate,
                         args=(mapping,), daemon=True).start()

    def _do_generate(self, mapping):
        try:
            df = pd.read_excel(self.data_file.get())
            total = len(df)
            os.makedirs(self.output_dir.get(), exist_ok=True)
            ok = 0

            self._log(f"[开始] 共 {total} 行数据，{len(mapping)} 条映射")
            self.progress.set(0)

            for i, (_, row) in enumerate(df.iterrows()):
                try:
                    wb = load_workbook(self.tmpl_file.get())
                    sname = self.sheet_name.get()
                    ws = wb[sname] if sname in wb.sheetnames else wb.active

                    for col_name, cell_addr in mapping:
                        if col_name in row.index:
                            val = row[col_name]
                            ws[cell_addr.strip().upper()] = None if pd.isna(val) else val

                    name_val = str(row.get(self.name_col.get(), f"row_{i+1}"))
                    # sanitize filename
                    for ch in r'\/:*?"<>|':
                        name_val = name_val.replace(ch, "_")
                    out_path = os.path.join(self.output_dir.get(), f"{name_val}.xlsx")
                    wb.save(out_path)
                    ok += 1
                    self._log(f"  ✓ [{i+1}/{total}] {name_val}.xlsx")
                except Exception as e:
                    self._log(f"  ✗ [{i+1}/{total}] 错误: {e}")

                pct = (i + 1) / total * 100
                self.progress.set(pct)
                self.status_var.set(f"进度 {i+1}/{total}")

            self._log(f"\n[完成] 成功生成 {ok}/{total} 个文件 → {self.output_dir.get()}")
            self.status_var.set(f"完成！成功 {ok}/{total}")
            messagebox.showinfo("完成", f"成功生成 {ok}/{total} 个 Excel 文件！\n\n输出目录：\n{self.output_dir.get()}")
        except Exception as e:
            self._log(f"[严重错误] {e}")
            messagebox.showerror("错误", str(e))

    # ─── Help Page ────────────────────────────────────────────────
    def _build_help_page(self):
        p = self.pages["help"]
        tk.Label(p, text="使用说明", bg=BG, fg=TEXT,
                 font=("Arial", 18, "bold")).pack(anchor="w", pady=(0, 16))
        card = self._card(p, "操作步骤")
        card.pack(fill=tk.BOTH, expand=True)
        card.configure(padx=20, pady=16)
        help_text = """
📋 使用步骤

① 文件配置
   • 选择「数据源文件」—— 每一行对应一个要生成的表单
   • 选择「模板文件」—— 格式固定的 Excel 模板
   • 选择「输出目录」—— 生成文件的保存位置
   • 设置「文件命名列」—— 用哪一列的值作为输出文件名
   • 点击「加载数据源列」读取列名

② 字段映射
   • 点击「＋ 添加映射行」
   • 左侧下拉选择数据源中的列名
   • 右侧输入模板中对应的单元格地址（如 B3、D5、F12）
   • 可保存配置，下次打开自动加载

③ 开始生成
   • 点击「刷新摘要」确认配置无误
   • 点击「开始批量生成」
   • 日志区域实时显示进度

─────────────────────────────────────────
⚠ 注意事项
  • 模板文件在生成过程中不会被修改，每行数据使用独立副本
  • 单元格地址不区分大小写（b3 与 B3 等效）
  • 文件名中的特殊字符（\\ / : * ? " < > |）会自动替换为下划线
  • 如数据源某列值为空，对应单元格将填入空值
"""
        txt = tk.Text(card, font=("Arial", 11), bg=CARD, fg=TEXT,
                      relief="flat", wrap="word", state="normal")
        txt.insert("1.0", help_text)
        txt.config(state="disabled")
        txt.pack(fill=tk.BOTH, expand=True)

    # ─── Config persistence ───────────────────────────────────────
    def _save_config(self):
        cfg = {
            "data_file":  self.data_file.get(),
            "tmpl_file":  self.tmpl_file.get(),
            "output_dir": self.output_dir.get(),
            "name_col":   self.name_col.get(),
            "sheet_name": self.sheet_name.get(),
            "mapping": [(c.get(), v.get()) for c, v in self.mapping_rows],
        }
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        messagebox.showinfo("保存", "配置已保存！")

    def _load_config(self):
        if not os.path.exists(CONFIG_FILE):
            return
        try:
            with open(CONFIG_FILE, encoding="utf-8") as f:
                cfg = json.load(f)
            self.data_file.set(cfg.get("data_file", ""))
            self.tmpl_file.set(cfg.get("tmpl_file", ""))
            self.output_dir.set(cfg.get("output_dir", ""))
            self.name_col.set(cfg.get("name_col", ""))
            self.sheet_name.set(cfg.get("sheet_name", "Sheet1"))
            if self.data_file.get() and os.path.exists(self.data_file.get()):
                try:
                    df = pd.read_excel(self.data_file.get(), nrows=0)
                    self.df_columns = list(df.columns)
                    self.name_col_combo["values"] = self.df_columns
                except Exception:
                    pass
            for child in self.map_frame.winfo_children():
                child.destroy()
            self.mapping_rows.clear()
            for col, cell in cfg.get("mapping", []):
                self._add_mapping_row(col, cell)
        except Exception:
            pass


if __name__ == "__main__":
    app = ExcelBatchApp()
    app.mainloop()
