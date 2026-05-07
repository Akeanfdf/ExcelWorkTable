from __future__ import annotations

import json
import os
from typing import Any, Dict, List, Tuple

import pandas as pd
from NodeGraphQt import BaseNode
from openpyxl import load_workbook

from excel_workflow.core.registry import register_node, register_runner
from excel_workflow.nodes.utils import parse_mapping_json, sanitize_filename


def _apply_mapping(ws, mapping: List[Tuple[str, str]], row: pd.Series) -> None:
    for col_name, cell_addr in mapping:
        if col_name in row.index:
            val = row[col_name]
            ws[cell_addr.strip().upper()] = None if pd.isna(val) else val


def _generate_from_df(
    df: pd.DataFrame,
    tmpl: str,
    out_dir: str,
    name_col: str,
    sheet_name: str,
    mapping: List[Tuple[str, str]],
    ctx,
    name_prefix: str = "",
) -> List[str]:
    os.makedirs(out_dir, exist_ok=True)
    if name_col and name_col not in df.columns:
        raise ValueError(f"命名列不存在: {name_col}")
    paths: List[str] = []
    total = len(df)
    for i, (_, row) in enumerate(df.iterrows()):
        wb = load_workbook(tmpl)
        sname = sheet_name or "Sheet1"
        ws = wb[sname] if sname in wb.sheetnames else wb.active
        _apply_mapping(ws, mapping, row)
        name_val = str(row.get(name_col, f"row_{i+1}")) if name_col else f"row_{i+1}"
        if name_prefix:
            name_val = f"{name_prefix}_{name_val}"
        name_val = sanitize_filename(name_val)
        out_path = os.path.join(out_dir, f"{name_val}.xlsx")
        wb.save(out_path)
        paths.append(out_path)
        ctx.log(f"  ✓ [{i+1}/{total}] {out_path}")
    return paths


def _pick_template(row: pd.Series, rules: List[dict], default: str) -> str:
    for r in rules:
        col = r.get("when_column") or r.get("column")
        eq = r.get("equals")
        tp = (r.get("template") or "").strip()
        if not col or not tp:
            continue
        if col not in row.index:
            continue
        if str(row.get(col, "")) == str(eq):
            return tp
    return default


@register_node
class TemplateMapper(BaseNode):
    __identifier__ = "excel_workflow.template"
    NODE_NAME = "模板填充"

    def __init__(self):
        super().__init__()
        self.set_color(239, 68, 68)
        self.add_input("dataframe")
        self.add_output("files")
        self.add_text_input("template_path", "模板 xlsx", tab="参数")
        self.add_text_input("output_dir", "输出目录", tab="参数")
        self.add_text_input("name_column", "文件命名列", tab="参数")
        self.add_text_input("sheet_name", "工作表名", tab="参数")
        self.add_text_input(
            "mapping_json",
            '映射 JSON [["列","单元格"],...]',
            tab="映射",
        )


@register_runner("excel_workflow.template.TemplateMapper")
def run_template_mapper(node, inputs, ctx) -> Dict[str, Any]:
    raw = inputs.get("dataframe")
    tmpl = (node.get_property("template_path") or "").strip()
    out_dir = (node.get_property("output_dir") or "").strip()
    name_col = (node.get_property("name_column") or "").strip()
    sheet_name = (node.get_property("sheet_name") or "Sheet1").strip() or "Sheet1"
    mapping = parse_mapping_json(node.get_property("mapping_json") or "[]")
    if not tmpl or not os.path.isfile(tmpl):
        raise FileNotFoundError("TemplateMapper: 模板文件无效")
    if not out_dir:
        raise ValueError("TemplateMapper: 未设置输出目录")
    if not mapping:
        raise ValueError("TemplateMapper: 映射为空")
    if not name_col:
        raise ValueError("TemplateMapper: 未设置命名列")

    all_files: List[str] = []
    if isinstance(raw, dict):
        ctx.log(f"TemplateMapper: 分组模式，共 {len(raw)} 组")
        for key, subdf in raw.items():
            ctx.log(f"  — 组 {key}: {len(subdf)} 行")
            prefix = sanitize_filename(str(key))
            all_files.extend(
                _generate_from_df(
                    subdf, tmpl, out_dir, name_col, sheet_name, mapping, ctx, prefix
                )
            )
    else:
        if raw is None:
            raise ValueError("TemplateMapper: 需要上游 dataframe")
        df = raw
        if not isinstance(df, pd.DataFrame):
            raise TypeError("TemplateMapper: 输入必须是 DataFrame 或分组字典")
        ctx.log(f"TemplateMapper: 共 {len(df)} 行")
        all_files = _generate_from_df(df, tmpl, out_dir, name_col, sheet_name, mapping, ctx)
    ctx.log(f"TemplateMapper: 完成，生成 {len(all_files)} 个文件")
    return {"files": all_files}


@register_node
class MultiTemplateMapper(BaseNode):
    __identifier__ = "excel_workflow.template"
    NODE_NAME = "多模板路由"

    def __init__(self):
        super().__init__()
        self.set_color(239, 68, 68)
        self.add_input("dataframe")
        self.add_output("files")
        self.add_text_input("output_dir", "输出目录", tab="参数")
        self.add_text_input("name_column", "文件命名列", tab="参数")
        self.add_text_input("sheet_name", "工作表名", tab="参数")
        self.add_text_input(
            "mapping_json",
            '映射 JSON [["列","单元格"],...]',
            tab="映射",
        )
        self.add_text_input(
            "rules_json",
            '[{"when_column":"类型","equals":"A","template":"C:/a.xlsx"}]',
            tab="规则",
        )
        self.add_text_input("default_template", "默认模板路径", tab="规则")


@register_runner("excel_workflow.template.MultiTemplateMapper")
def run_multi_template(node, inputs, ctx) -> Dict[str, Any]:
    raw = inputs.get("dataframe")
    out_dir = (node.get_property("output_dir") or "").strip()
    name_col = (node.get_property("name_column") or "").strip()
    sheet_name = (node.get_property("sheet_name") or "Sheet1").strip() or "Sheet1"
    mapping = parse_mapping_json(node.get_property("mapping_json") or "[]")
    rules_raw = (node.get_property("rules_json") or "[]").strip() or "[]"
    rules = json.loads(rules_raw)
    if not isinstance(rules, list):
        raise ValueError("rules_json 必须是数组")
    default_t = (node.get_property("default_template") or "").strip()
    if not out_dir:
        raise ValueError("MultiTemplateMapper: 输出目录未设置")
    if not mapping:
        raise ValueError("MultiTemplateMapper: 映射为空")
    if not name_col:
        raise ValueError("MultiTemplateMapper: 命名列未设置")

    def process_df(df: pd.DataFrame, prefix: str = "") -> List[str]:
        files: List[str] = []
        total = len(df)
        for i, (_, row) in enumerate(df.iterrows()):
            tmpl = _pick_template(row, rules, default_t)
            if not tmpl or not os.path.isfile(tmpl):
                raise FileNotFoundError(f"行 {i+1}: 模板无效 ({tmpl})")
            wb = load_workbook(tmpl)
            sname = sheet_name or "Sheet1"
            ws = wb[sname] if sname in wb.sheetnames else wb.active
            _apply_mapping(ws, mapping, row)
            name_val = str(row.get(name_col, f"row_{i+1}"))
            if prefix:
                name_val = f"{prefix}_{name_val}"
            name_val = sanitize_filename(name_val)
            os.makedirs(out_dir, exist_ok=True)
            out_path = os.path.join(out_dir, f"{name_val}.xlsx")
            wb.save(out_path)
            files.append(out_path)
            ctx.log(f"  ✓ [{i+1}/{total}] {out_path}")
        return files

    all_files: List[str] = []
    if isinstance(raw, dict):
        for key, subdf in raw.items():
            pre = sanitize_filename(str(key))
            all_files.extend(process_df(subdf, pre))
    else:
        if raw is None:
            raise ValueError("需要上游 dataframe")
        all_files = process_df(raw)
    return {"files": all_files}
